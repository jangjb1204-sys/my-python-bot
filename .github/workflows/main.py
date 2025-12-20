import pandas as pd
pd.set_option('future.no_silent_downcasting', True)
import yfinance as yf
from datetime import datetime, timedelta
import requests
import json
import numpy as np
import matplotlib.pyplot as plt
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import time  # 시간 지연을 위해 추가
import os

# 1. 현재 실행 중인 파이썬 스크립트의 경로를 가져옵니다.
current_dir = os.getcwd()

# 2. 저장할 하위 폴더명을 설정합니다.
folder_name = "US_stock"
folder_path = os.path.join(current_dir, folder_name)

# 3. 폴더가 없다면 생성합니다 (exist_ok=True는 이미 폴더가 있어도 에러를 내지 않습니다).
os.makedirs(folder_path, exist_ok=True)


# --- 상수 정의 ---
SEARCH_DAYS = 365 * 4
DATE_FORMAT = '%Y-%m-%d'
COLORS = {
    'GREEN': '#008000',
    'RED': '#FF0000',
    'BLACK': '#000000',
    'YELLOW_FILL': '#FFFF99',
    'HEADER_FILL': '#D3D3D3',
    'LATEST_FILL': '#E6E6FA',
}
USER_AGENT = (
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
    '(KHTML, like Gecko) Chrome/89.0.4389.82 Safari/537.36'
)

# --- 데이터 수집 함수 (공통 데이터용) ---
def fetch_fear_and_greed_index(start_date: str) -> pd.DataFrame | None:
    time.sleep(1) # API 매너 호출
    url = f'https://production.dataviz.cnn.io/index/fearandgreed/graphdata/{start_date}'
    headers = {'User-Agent': USER_AGENT}
    try:
        response = requests.get(url, headers=headers, timeout=10)
        if response.status_code != 200:
            print(f"F&G 요청 실패: {response.status_code}")
            return None
        data = json.loads(response.text)
        data_list = data['fear_and_greed_historical']['data']
        df = pd.DataFrame([
            {
                'Date': datetime.fromtimestamp(item['x'] / 1000).strftime(DATE_FORMAT),
                'FG index': round(item['y']),
                'rating': item.get('rating', 'N/A')
            }
            for item in data_list
        ])
        df['Date'] = pd.to_datetime(df['Date'])
        df = df.sort_values('Date').drop_duplicates('Date', keep='first')
        print("Fear and Greed 데이터 수집 완료")
        return df
    except Exception as e:
        print(f"F&G 데이터 오류: {e}")
        return None

def fetch_common_market_data(period: str):
    """
    모든 종목에 공통으로 쓰이는 지표들을 한 번에 수집합니다.
    """
    print("공통 시장 지표(VIX, 국채금리 등) 수집 중...")
    
    # 1. 10년물 국채
    try:
        treasury = yf.Ticker('^TNX').history(period=period)[['Close']].rename(columns={'Close': '10Y Treasury'})
        treasury = treasury.reset_index()
        treasury['Date'] = pd.to_datetime(treasury['Date'].dt.date)
        treasury['10Y Treasury'] = treasury['10Y Treasury'].round(2)
        time.sleep(1)
    except Exception:
        treasury = pd.DataFrame()

    # 2. VIX
    try:
        vix = yf.Ticker('^VIX').history(period=period)[['Close']].rename(columns={'Close': 'VIX'})
        vix = vix.reset_index()
        vix['Date'] = pd.to_datetime(vix['Date'].dt.date)
        vix['VIX'] = vix['VIX'].round(1)
        time.sleep(1)
    except Exception:
        vix = pd.DataFrame()

    # 3. VIX1D
    try:
        vix1d = yf.Ticker('^VIX1D').history(period=period)[['Close']].rename(columns={'Close': 'VIX1D'})
        vix1d = vix1d.reset_index()
        vix1d['Date'] = pd.to_datetime(vix1d['Date'].dt.date)
        vix1d['VIX1D'] = vix1d['VIX1D'].round(1)
        time.sleep(1)
    except Exception:
        vix1d = pd.DataFrame()

    # 4. SKEW
    try:
        skew = yf.Ticker('^SKEW').history(period=period)[['Close']].rename(columns={'Close': 'SKEW'})
        skew = skew.reset_index()
        skew['Date'] = pd.to_datetime(skew['Date'].dt.date)
        skew['SKEW'] = skew['SKEW'].round(1)
        time.sleep(1)
    except Exception:
        skew = pd.DataFrame()
        
    # 5. Fear and Greed
    start_date = (datetime.now() - timedelta(days=SEARCH_DAYS)).strftime(DATE_FORMAT)
    fg_data = fetch_fear_and_greed_index(start_date)

    return {
        'treasury': treasury,
        'vix': vix,
        'vix1d': vix1d,
        'skew': skew,
        'fg_data': fg_data
    }

def fetch_stock_data(ticker: str, period: str) -> pd.DataFrame:
    # 종목 데이터 요청 시 약간의 딜레이 추가
    time.sleep(2) 
    stock = yf.Ticker(ticker)
    data = stock.history(period=period).reset_index()
    if not data.empty:
        data['Date'] = pd.to_datetime(data['Date'].dt.date)
    return data

# --- 기술적 지표 계산 함수 (기존 유지) ---
def calculate_rsi(data: pd.DataFrame, window: int = 14) -> pd.Series:
    if len(data) < window:
        return pd.Series([np.nan] * len(data), index=data.index)
    delta = data['Close'].diff()
    gain = delta.where(delta > 0, 0).rolling(window=window).mean()
    loss = (-delta.where(delta < 0, 0)).rolling(window=window).mean()
    rs = gain / loss
    return (100 - (100 / (1 + rs))).round(2)

def calculate_stochastic_slow(data: pd.DataFrame, n: int = 14, m: int = 3, t: int = 3) -> tuple:
    if len(data) < n:
        return pd.Series([np.nan] * len(data), index=data.index), pd.Series([np.nan] * len(data), index=data.index)
    low_min = data['Low'].rolling(window=n).min()
    high_max = data['High'].rolling(window=n).max()
    k_fast = 100 * ((data['Close'] - low_min) / (high_max - low_min))
    slow_k = k_fast.rolling(window=m).mean().round(2)
    slow_d = slow_k.rolling(window=t).mean().round(2)
    return slow_k, slow_d

def calculate_moving_averages(data: pd.DataFrame, windows: list = [20, 60, 120, 200]) -> pd.DataFrame:
    for window in windows:
        if len(data) >= window:
            data[f'MA{window}'] = data['Close'].rolling(window=window).mean().round(2)
        else:
            data[f'MA{window}'] = np.nan
    return data

# --- 매매 신호 생성 함수 (기존 유지) ---
def generate_stochastic_signals(data: pd.DataFrame) -> pd.DataFrame:
    data['SS Signal'] = ''
    if 'Slow_K' in data.columns and 'Slow_D' in data.columns:
        data.loc[(data['Slow_K'].shift(1) < data['Slow_D'].shift(1)) & (data['Slow_K'] > data['Slow_D']), 'SS Signal'] = 'Buy'
        data.loc[(data['Slow_K'].shift(1) > data['Slow_D'].shift(1)) & (data['Slow_K'] < data['Slow_D']), 'SS Signal'] = 'Sell'
        data['SS value'] = np.where(data['SS Signal'] != '', data['Slow_K'], '')
    return data

def generate_fg_rsi_signals(data: pd.DataFrame) -> pd.DataFrame:
    def apply_rules(row):
        # RSI가 NaN이면 신호 없음 처리
        if pd.isna(row.get('RSI')): 
            return ''
        
        rsi = row['RSI']
        fg_idx = row.get('FG index', -1) # FG 없을 경우 대비

        # FG 데이터가 유효한지 확인 (None이나 NaN이 아닐 때)
        has_fg = not pd.isna(fg_idx) and fg_idx != -1

        if rsi >= 60 or (has_fg and 51 <= fg_idx <= 100):
            return 'BUY STOP'
        elif rsi <= 30 or (has_fg and 26 <= fg_idx <= 50):
            return '2x BUY'
        elif rsi <= 20 or (has_fg and 0 <= fg_idx <= 25):
            return '3x BUY'
        return '1x BUY'
        
    data['FG/RSI signal'] = data.apply(apply_rules, axis=1)
    return data

def generate_puddle_signals(data: pd.DataFrame) -> pd.DataFrame:
    alerts = ['']
    for i in range(1, len(data)):
        row, prev = data.iloc[i], data.iloc[i-1]
        conditions = {
            1: (not pd.isna(row['MA20']) and row['Close'] < row['MA20'] and prev['Close'] >= prev['MA20']),
            2: (not pd.isna(row['MA60']) and row['Close'] < row['MA60'] and prev['Close'] >= prev['MA60']),
            3: (not pd.isna(row['MA120']) and row['Close'] < row['MA120'] and prev['Close'] >= prev['MA120']),
            4: (not pd.isna(row['MA200']) and row['Close'] < row['MA200'] and not pd.isna(row['RSI']) and row['RSI'] < 30)
        }
        timings = [k for k, v in conditions.items() if v]
        alerts.append({
            4: '4th: MA200, RSI≤30, 100% cash, 40d',
            3: '3rd: MA120, 50% cash, 5d',
            2: '2nd: MA60, 50% cash, 5d',
            1: '1st: MA20, 10% cash'
        }.get(max(timings)) if timings else '')
    data['Puddle'] = alerts
    return data

def calculate_vix_skew_signals(data: pd.DataFrame) -> pd.DataFrame:
    if 'VIX' in data.columns and 'VIX1D' in data.columns:
        data['VIX1D>VIX'] = np.where(
            (data['VIX'].notna()) & (data['VIX1D'].notna()) & 
            (data['VIX'] >= 25) & (data['VIX1D'] > data['VIX']), 
            'BUY', ''
        )
    else:
        data['VIX1D>VIX'] = ''
    return data

# --- 데이터 처리 메인 함수 (구조 변경) ---
def process_stock_data(ticker: str, name: str, common_data: dict, period: str = '2y', delta: int = 400) -> pd.DataFrame:
    """
    common_data: 미리 수집한 공통 지표 딕셔너리
    """
    data = fetch_stock_data(ticker, period)
    if data.empty:
        print(f"{ticker} - No stock data available.")
        return pd.DataFrame()
    print(f"{ticker} - Initial data rows: {len(data)}")

    # 공통 데이터 가져오기
    fg_data = common_data.get('fg_data')
    treasury_data = common_data.get('treasury')
    vix_data = common_data.get('vix')
    vix1d_data = common_data.get('vix1d')
    skew_data = common_data.get('skew')

    data[['Close', 'Open', 'High', 'Low']] = data[['Close', 'Open', 'High', 'Low']].round(2)
    data['Change(%)'] = (data['Close'].pct_change() * 100).round(2)
    log_returns = np.log(data['Close'] / data['Close'].shift(1))
    data['2sigma(%)'] = round(log_returns.std() * 100 * 2, 1)

    data = calculate_moving_averages(data)
    data['RSI'] = calculate_rsi(data)
    data['Slow_K'], data['Slow_D'] = calculate_stochastic_slow(data)

    # 공통 데이터 병합 (데이터가 존재할 경우에만)
    if fg_data is not None and not fg_data.empty:
        data = pd.merge(data, fg_data, on='Date', how='left')
    if treasury_data is not None and not treasury_data.empty:
        data = pd.merge(data, treasury_data, on='Date', how='left')
    if vix_data is not None and not vix_data.empty:
        data = pd.merge(data, vix_data, on='Date', how='left')
    if vix1d_data is not None and not vix1d_data.empty:
        data = pd.merge(data, vix1d_data, on='Date', how='left')
    if skew_data is not None and not skew_data.empty:
        data = pd.merge(data, skew_data, on='Date', how='left')

    data = generate_stochastic_signals(data)
    data = generate_fg_rsi_signals(data)
    data = generate_puddle_signals(data)
    data = calculate_vix_skew_signals(data)

    data = data[data['Date'] >= (datetime.now() - timedelta(days=delta))] if len(data) > delta else data
    print(f"{ticker} - Final data rows: {len(data)}")

    data['Tick'] = ticker
    
    # 필요한 컬럼만 선택 (컬럼이 없을 경우를 대비해 reindex 사용)
    target_columns = [
        'Tick', 'Date', 'Open', 'High', 'Low', 'Close', 'Volume', 'Change(%)', '2sigma(%)',
        'MA20', 'MA60', 'MA120', 'MA200', 'RSI', 'FG index', 'rating',
        'FG/RSI signal', 'Puddle', '10Y Treasury', 'VIX', 'VIX1D', 'VIX1D>VIX', 'SKEW'
    ]
    # 실제 존재하는 컬럼만 남기고, 없는건 빈값으로 채우기 위해 로직 보강
    existing_cols = [c for c in target_columns if c in data.columns]
    data_output = data[existing_cols].copy()
    
    # 누락된 컬럼 추가
    for col in target_columns:
        if col not in data_output.columns:
            data_output[col] = ''
            
    data_output = data_output[target_columns].fillna('') # 순서 정렬 및 NaN 처리

    if not data_output.empty:
        create_table_image(data_output, name)
        save_to_excel(data_output, name)
    return data_output

# --- 시각화 및 저장 함수 (기존 유지) ---
def create_table_image(data: pd.DataFrame, name: str) -> None:
    columns_to_drop = ['MA20', 'MA60', 'MA120', 'MA200', 'Open', 'High', 'Low', 'Volume']
    # 없는 컬럼은 드랍하지 않도록 필터링
    cols_to_drop_actual = [c for c in columns_to_drop if c in data.columns]
    
    last_n = data.tail(20).drop(columns=cols_to_drop_actual, axis=1)
    if last_n.empty:
        return
    
    last_n['Date'] = last_n['Date'].dt.strftime(DATE_FORMAT)
    latest_date = last_n['Date'].iloc[-1] if not last_n.empty else ''
    last_n = last_n.fillna('')

    fig, ax = plt.subplots(figsize=(16, max(3, len(last_n) * 0.2)))
    ax.axis('tight')
    ax.axis('off')
    table = ax.table(cellText=last_n.values, colLabels=last_n.columns, loc='center', cellLoc='center')
    table.auto_set_font_size(False)
    table.set_fontsize(10)

    column_widths = [max(last_n[col].astype(str).str.len().max(), len(col)) * 0.01 for col in last_n.columns]
    for i, width in enumerate(column_widths):
        table.auto_set_column_width(i)
        table[0, i].set_width(width)
    table.scale(1.1, 1.5)

    # 인덱스 안전하게 찾기
    def get_col_idx(col_name):
        return last_n.columns.get_loc(col_name) if col_name in last_n.columns else -1

    change_col_idx = get_col_idx('Change(%)')
    rsi_col_idx = get_col_idx('RSI')
    vix_col_idx = get_col_idx('VIX')
    vix1d_col_idx = get_col_idx('VIX1D')
    vix1d_gt_vix_col_idx = get_col_idx('VIX1D>VIX')
    skew_col_idx = get_col_idx('SKEW')

    for (row_idx, col_idx), cell in table.get_celld().items():
        data_idx = row_idx - 1
        if row_idx == 0:
            cell.set_facecolor(COLORS['HEADER_FILL'])
            cell.set_text_props(weight='bold', color=COLORS['BLACK'])
        elif data_idx >= 0:
            # 안전하게 값 가져오기
            val_change = last_n['Change(%)'].iloc[data_idx] if 'Change(%)' in last_n.columns else ''
            
            if col_idx == change_col_idx and val_change != '':
                change = float(val_change)
                color = COLORS['RED'] if change < 0 else COLORS['GREEN'] if change > 0 else COLORS['BLACK']
                cell.set_text_props(color=color)
            elif col_idx == rsi_col_idx and last_n['RSI'].iloc[data_idx] != '':
                rsi = float(last_n['RSI'].iloc[data_idx])
                color = COLORS['GREEN'] if rsi <= 30 else COLORS['RED'] if rsi >= 70 else COLORS['BLACK']
                cell.set_text_props(color=color)
            elif col_idx == vix_col_idx and last_n['VIX'].iloc[data_idx] != '':
                vix = float(last_n['VIX'].iloc[data_idx])
                color = COLORS['GREEN'] if vix > 25 else COLORS['BLACK']
                cell.set_text_props(color=color)
            elif col_idx == vix1d_col_idx and last_n['VIX1D'].iloc[data_idx] != '':
                vix1d = float(last_n['VIX1D'].iloc[data_idx])
                color = COLORS['GREEN'] if vix1d > 25 else COLORS['BLACK']
                cell.set_text_props(color=color)
            elif col_idx == vix1d_gt_vix_col_idx and last_n['VIX1D>VIX'].iloc[data_idx] == 'BUY':
                cell.set_text_props(color=COLORS['GREEN'])
            elif col_idx == skew_col_idx and last_n['SKEW'].iloc[data_idx] != '':
                skew = float(last_n['SKEW'].iloc[data_idx])
                color = COLORS['RED'] if skew >= 155 else COLORS['GREEN'] if skew <= 127 else COLORS['BLACK']
                cell.set_text_props(color=color)
            
            # 배경색 로직
            if last_n['Date'].iloc[data_idx] == latest_date:
                cell.set_facecolor(COLORS['LATEST_FILL'])
            elif (val_change != '' and '2sigma(%)' in last_n.columns and last_n['2sigma(%)'].iloc[data_idx] != '' and
                  float(val_change) < -float(last_n['2sigma(%)'].iloc[data_idx])):
                cell.set_facecolor(COLORS['YELLOW_FILL'])
        cell.set_edgecolor(COLORS['BLACK'])
        cell.set_linewidth(0.25)

    # 경로 수정 필요시 변경
    save_path = os.path.join(folder_path, f"{name}_table.jpg")
    
    plt.savefig(save_path, bbox_inches='tight', dpi=300)
    plt.close()

def create_puddle_trading_chart(data: pd.DataFrame, name: str) -> None:
    if data.empty:
        return
    
    fig = make_subplots(rows=1, cols=1)
    fig.add_trace(go.Candlestick(
        x=data['Date'], open=data['Open'], high=data['High'], low=data['Low'], close=data['Close'],
        name='Price', increasing_line_color='#26A69A', decreasing_line_color='#EF5350',
        increasing_fillcolor='#26A69A', decreasing_fillcolor='#EF5350'
    ))

    ma_colors = {'MA20': '#2196F3', 'MA60': '#FF9800', 'MA120': '#F44336', 'MA200': '#9C27B0'}
    for ma, color in ma_colors.items():
        if ma in data.columns and data[ma].notna().any():
            fig.add_trace(go.Scatter(x=data['Date'], y=data[ma], name=ma, line=dict(color=color, width=2), mode='lines'))

    if 'Puddle' in data.columns:
        buy_signals = data[data['Puddle'].str.contains(r'[a-zA-Z]', na=False)]['Date']
        if not buy_signals.empty:
            fig.add_trace(go.Scatter(
                x=buy_signals, y=data.loc[data['Date'].isin(buy_signals), 'Low'] * 0.985,
                mode='markers', name='Buy Signal', marker=dict(symbol='triangle-up', size=12, color='#FF1744', line=dict(width=2))
            ))

    fig.update_layout(
        title=dict(text=f'({name}) 차트', x=0.5, y=0.95, font=dict(size=20, color='#333')),
        plot_bgcolor='white', paper_bgcolor='white', font=dict(family='Arial', size=12, color='#444'),
        showlegend=True, legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5),
        xaxis_rangeslider_visible=False, height=600, margin=dict(l=80, r=20, t=85, b=50)
    )

    tick_dates = data['Date'][::max(1, len(data)//10)]
    fig.update_xaxes(
        tickmode='array', tickvals=tick_dates, ticktext=tick_dates.dt.strftime(DATE_FORMAT),
        tickangle=-90, showgrid=True, gridwidth=1, gridcolor='#f0f0f0', zeroline=False
    )
    fig.update_yaxes(
        tickformat=',d', showgrid=True, gridwidth=1, gridcolor='#f0f0f0', zeroline=False,
        title_text='Price', automargin=True, ticks='outside', ticklabelposition='outside',
        side='left', mirror=True
    )
    
    save_path = save_path = os.path.join(folder_path, f"{name}_chart.jpg")
    fig.write_image(save_path, width=800, height=500, scale=2)

def save_to_excel(data: pd.DataFrame, name: str) -> None:
    if data.empty:
        return
    
    data_for_excel = data.copy()
    data_for_excel['Date'] = data_for_excel['Date'].dt.strftime(DATE_FORMAT)
    data_for_excel = data_for_excel.fillna('')
    excel_file = save_path = os.path.join(folder_path, f"{name}_table.xlsx")
    data_for_excel.to_excel(excel_file, index=False, engine='openpyxl')

    wb = Workbook()
    ws = wb.active
    df = pd.read_excel(excel_file)

    for c_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=c_idx, value=col_name)
    for r_idx, row in enumerate(df.values, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx + 1, column=c_idx, value=value if pd.notna(value) else '')

    yellow_fill = PatternFill(start_color=COLORS['YELLOW_FILL'].lstrip('#'), 
                              end_color=COLORS['YELLOW_FILL'].lstrip('#'), 
                              fill_type='solid')
    red_font = Font(color=COLORS['RED'].lstrip('#'))
    green_font = Font(color=COLORS['GREEN'].lstrip('#'))

    # 인덱스 계산 (엑셀은 1부터 시작)
    def get_excel_col_idx(name):
        return df.columns.get_loc(name) + 1 if name in df.columns else -1
        
    change_col_idx = get_excel_col_idx('Change(%)')
    rsi_col_idx = get_excel_col_idx('RSI')
    vix_col_idx = get_excel_col_idx('VIX')
    vix1d_col_idx = get_excel_col_idx('VIX1D')
    vix1d_gt_vix_col_idx = get_excel_col_idx('VIX1D>VIX')
    skew_col_idx = get_excel_col_idx('SKEW')

    for row_idx, row in df.iterrows():
        # 2sigma 조건부 서식
        if ('Change(%)' in row and '2sigma(%)' in row and 
            row['Change(%)'] != '' and row['2sigma(%)'] != '' and 
            float(row['Change(%)']) < -float(row['2sigma(%)'])):
            for col_idx in range(1, len(df.columns) + 1):
                ws.cell(row=row_idx + 2, column=col_idx).fill = yellow_fill
                
        # 텍스트 색상
        if change_col_idx != -1 and row['Change(%)'] != '':
            change = float(row['Change(%)'])
            cell = ws.cell(row=row_idx + 2, column=change_col_idx)
            cell.font = red_font if change < 0 else green_font if change > 0 else Font()
            
        if rsi_col_idx != -1 and row['RSI'] != '':
            rsi = float(row['RSI'])
            cell = ws.cell(row=row_idx + 2, column=rsi_col_idx)
            cell.font = green_font if rsi <= 30 else red_font if rsi >= 70 else Font()
            
        if vix_col_idx != -1 and row['VIX'] != '':
            vix = float(row['VIX'])
            cell = ws.cell(row=row_idx + 2, column=vix_col_idx)
            cell.font = green_font if vix > 25 else Font()
            
        if vix1d_col_idx != -1 and row['VIX1D'] != '':
            vix1d = float(row['VIX1D'])
            cell = ws.cell(row=row_idx + 2, column=vix1d_col_idx)
            cell.font = green_font if vix1d > 25 else Font()
            
        if vix1d_gt_vix_col_idx != -1 and row['VIX1D>VIX'] == 'BUY':
            cell = ws.cell(row=row_idx + 2, column=vix1d_gt_vix_col_idx)
            cell.font = green_font
            
        if skew_col_idx != -1 and row['SKEW'] != '':
            skew = float(row['SKEW'])
            cell = ws.cell(row=row_idx + 2, column=skew_col_idx)
            cell.font = red_font if skew >= 155 else green_font if skew <= 127 else Font()

    for col_idx, column in enumerate(df.columns, 1):
        # 길이 계산 시 None 타입 에러 방지
        max_length = max(len(str(ws[f"{get_column_letter(col_idx)}{row_idx + 1}"].value or '')) 
                         for row_idx in range(len(df) + 1))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    wb.save(excel_file)

# --- 설정 및 실행 ---
def get_ticker_configs() -> dict:
    return {
        'SOXL': 'SOXL',
        '^GSPC': 'S&P500',
        '^IXIC': 'NASDAQ',
        'SSO': 'SSO',
        'QLD': 'QLD',
        'GLD': 'GOLD',
        'FINX' : 'FINX',
        'BTGD' : 'BTGD',
        'SLV' : 'SILVER'
    }

if __name__ == '__main__':
    ticker_configs = get_ticker_configs()
    
    # 1. 공통 데이터 먼저 수집 (루프 밖에서 한 번만!)
    common_market_data = fetch_common_market_data(period='4y')
    
    for ticker, name in ticker_configs.items():
        print(f"Processing {ticker} ({name})...")
        try:
            # 2. 공통 데이터를 함수에 전달
            data = process_stock_data(ticker, name, common_market_data, period='4y', delta=600)
            if not data.empty:
                create_puddle_trading_chart(data, name)
            print(f"Completed {ticker} ({name}).")
            
            # 3. 종목 간 처리 딜레이 추가
            time.sleep(3) 
            
        except Exception as e:
            print(f"Error processing {ticker} ({name}): {str(e)}")
            
            
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import os
from datetime import datetime

# --- 설정 ---
BASE_DIR = "US_stock"
OUTPUT_DIR_A = BASE_DIR
OUTPUT_DIR_B = BASE_DIR

# --- 그래프 설정 ---
GRAPH_COLORS = {
    'Close': '#1F77B4',
    'VIX1D>VIX': '#FF4500',
    'Puddle': '#32CD32',
    'YearGrid': '#A9A9A9',
    'Skew': '#FF0000'
}
GRAPH_MARKERS = {
    'Skew': '^'
}
GRAPH_STYLES = {
    'Close': {'linewidth': 2.5},
    'VIX1D>VIX': {'linestyle': '-', 'linewidth': 2, 'alpha': 0.5},
    'Puddle': {'s': 100, 'edgecolors': 'black', 'linewidth': 0.5, 'alpha': 0.5, 'zorder': 5},
    'Skew': {'s': 100, 'edgecolors': 'red', 'linewidth': 0.5, 'alpha': 0.5, 'zorder': 5}
}

# --- 데이터 로드 ---
def load_data(file_path: str) -> pd.DataFrame:
    try:
        df = pd.read_excel(file_path)
        df['Date'] = pd.to_datetime(df['Date'])
        return df
    except FileNotFoundError:
        print(f"Error: File not found at {file_path}")
        return pd.DataFrame()
    except pd.errors.ParserError:
        print(f"Error: Invalid Excel file format at {file_path}")
        return pd.DataFrame()
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return pd.DataFrame()

# --- 날짜 범위에 따른 눈금 설정 ---
def set_date_ticks(ax, date_range_years):
    if date_range_years <= 2:
        ax.xaxis.set_major_locator(mdates.MonthLocator(interval=1))
    else:
        ax.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))

# --- 라인 그래프 시각화 함수 (ax2 제거) ---
def plot_line_graph_with_rsi(df: pd.DataFrame, output_path: str, title: str, y_label: str) -> None:
    required_cols = ['Date', 'Close', 'VIX1D>VIX', 'Puddle', 'SKEW', 'RSI']
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        print(f"Missing columns: {missing_cols}")
        if 'RSI' in missing_cols:
            print("RSI column is required for this visualization.")
        return

    df['Close'] = pd.to_numeric(df['Close'], errors='coerce')
    df['RSI'] = pd.to_numeric(df['RSI'], errors='coerce')

    min_date = df['Date'].min()
    max_date = df['Date'].max()
    date_range_years = (max_date - min_date).days / 365

    # 단일 플롯 생성 (ax2 제거)
    fig, ax1 = plt.subplots(figsize=(10, 6))
    
    # 주가 그래프
    ax1.plot(df['Date'], df['Close'], label=f'{title} Close', 
             color='black', linewidth=2.5)

    # 모든 시그널에 동일한 크기와 투명도 적용을 위한 공통 스타일
    common_signal_style = {'s': 80, 'alpha': 0.5, 'edgecolors': 'black', 'linewidth': 0.5, 'zorder': 5}

    # 신호 표시
    buy_signals = df[df['VIX1D>VIX'] == 'BUY']
    if not buy_signals.empty:
        for date in buy_signals['Date']:
            ax1.axvline(x=date, color=GRAPH_COLORS['VIX1D>VIX'], 
                        label='VIX1D>VIX Buy' if date == buy_signals['Date'].iloc[0] else "", 
                        **GRAPH_STYLES['VIX1D>VIX'])

    skew_signals = df[df['SKEW'] <= 130]
    #체크#if not skew_signals.empty:
    #체크#    ax1.scatter(skew_signals['Date'], skew_signals['Close'], 
    #체크#               label='Skew <= 130', color=GRAPH_COLORS['Skew'], 
    #체크#               marker=GRAPH_MARKERS['Skew'],  **{**common_signal_style, 'edgecolors': 'red', 'zorder': 6})

    puddle_signals = df[df['Puddle'].str.contains(r'[a-zA-Z]', na=False)]
    #체크#if not puddle_signals.empty:
    #체크#    ax1.scatter(puddle_signals['Date'], puddle_signals['Close'], 
    #체크#               label='Puddle Signal', color=GRAPH_COLORS['Puddle'], **common_signal_style)

    # 연도별 그리드 라인
    for year in df['Date'].dt.year.unique():
        year_date = datetime(year, 1, 1)
        if year_date >= min_date and year_date <= max_date:
            ax1.axvline(x=year_date, color=GRAPH_COLORS['YearGrid'], linestyle=':', 
                        linewidth=2, alpha=0.3)

    # RSI 기반 신호 (ax1에 표시)
    oversold_points = df[df['RSI'] <= 30]
    #체크#if not oversold_points.empty:
    #체크#    ax1.scatter(oversold_points['Date'], oversold_points['Close'], 
    #체크#               label='RSI Oversold', color='#0066CC', marker='o', **common_signal_style)
    
    overbought_points = df[df['RSI'] >= 70]
    #if not overbought_points.empty:
    #    ax1.scatter(overbought_points['Date'], overbought_points['Close'], 
    #               label='RSI Overbought', color='red', marker='o', **common_signal_style)

    # RSI와 Puddle 겹치는 지점
    overlap_points = pd.merge(oversold_points, puddle_signals, on='Date', how='inner')
    if not overlap_points.empty:
        ax1.scatter(overlap_points['Date'], overlap_points['Close_x'], 
                    label='RSI & Puddle Overlap', color='purple', **common_signal_style)

    #체크#non_overlap_oversold = oversold_points[~oversold_points['Date'].isin(overlap_points['Date'])]
    #체크#if not non_overlap_oversold.empty:
    #체크#    ax1.scatter(non_overlap_oversold['Date'], non_overlap_oversold['Close'], 
    #체크#               label='_nolegend_', color='#0066CC', marker='o', **common_signal_style)

    #체크#non_overlap_puddle = puddle_signals[~puddle_signals['Date'].isin(overlap_points['Date'])]
    #체크#if not non_overlap_puddle.empty:
    #체크#    ax1.scatter(non_overlap_puddle['Date'], non_overlap_puddle['Close'], 
    #체크#               label='_nolegend_', color=GRAPH_COLORS['Puddle'], **common_signal_style)

    # 그래프 스타일 설정
    set_date_ticks(ax1, date_range_years)
    ax1.set_xlim(min_date, max_date)
    
    ax1.tick_params(axis='x', rotation=80, labelsize=10)
    ax1.set_ylabel(y_label, fontsize=12)
    ax1.set_xlabel('DATE', fontsize=12)
    ax1.tick_params(axis='y', labelsize=10)
    
    ax1.set_facecolor('#FAFAFA')
    fig.patch.set_facecolor('#FFFFFF')
    
    ax1.set_title(f'{title} WITH SIGNAL', fontsize=16, pad=10)
    ax1.legend(loc='upper left', fontsize=9, frameon=True, edgecolor='black', 
               facecolor='white', framealpha=1)
    
    
    # x축 범위 확장 및 subplot 여백 조정
    padding_ratio = 0.02  # 데이터 범위의 5%만큼 여백 추가
    x_min = min_date - pd.Timedelta(days=(max_date - min_date).days * padding_ratio)
    x_max = max_date + pd.Timedelta(days=(max_date - min_date).days * padding_ratio)
    ax1.set_xlim(x_min, x_max)
    
    plt.subplots_adjust(left=0.1, right=0.9)  # 좌우 subplot 여백 설정


    plt.tight_layout()
    
    plt.savefig(os.path.join(output_path, f'{title}_with signals.png'), 
                dpi=300, bbox_inches='tight')
    plt.close()

# --- 메인 실행 ---
def main():
    # OUTPUT_DIR_A에서 모든 .xlsx 파일 가져오기
    excel_files = [f for f in os.listdir(OUTPUT_DIR_A) if f.endswith('.xlsx')]
    
    if not excel_files:
        print("No Excel files found in the specified directory.")
        return

    # 각 파일에 대해 그래프 생성
    for file_name in excel_files:
        selected_file = os.path.join(OUTPUT_DIR_A, file_name)
        output_file_name = os.path.splitext(file_name)[0].replace('_table', '')

        # 데이터 로드
        df = load_data(selected_file)
        if df.empty:
            print(f"No data to analyze for {file_name}. Skipping...")
            continue

        # 그래프 생성
        plot_line_graph_with_rsi(df, OUTPUT_DIR_B, output_file_name, f'{output_file_name} VALUE')
        print(f"Graph saved for {file_name} at {OUTPUT_DIR_B}/{output_file_name}_with_rsi_signals.png")

if __name__ == "__main__":
    main()
